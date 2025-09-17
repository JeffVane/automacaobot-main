from flask import Flask, render_template, request, redirect, url_for, session, jsonify,Response
from flask_cors import CORS
import json
from typing import Optional
import threading
import time
# Certifique-se de que estes módulos existem e estão acessíveis
from maps_search import buscar_dados_cards_maps
from search_engine import buscar_links_site_maps
from scraper import extrair_contatos
import os, requests
from pprint import pprint # <<<< Mude para ESTA linha
from pathlib import Path
from dotenv import load_dotenv,find_dotenv
import re
import unicodedata
import io, csv
from datetime import datetime, date, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
WHATSAPP_TEXT_LIMIT = 4096  # limite seguro para texto puro
# tenta achar um .env a partir do cwd; se não achar, usa o .env ao lado deste arquivo
dotenv_path = find_dotenv(usecwd=True) or str(Path(__file__).resolve().parent / ".env")
print("DEBUG DOTENV_PATH:", dotenv_path)
load_dotenv(dotenv_path=dotenv_path, override=True)


def to_button_id(title: str) -> str:
    s = unicodedata.normalize("NFKD", title or "")
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s or "btn"

# O que vier do template (texto/payload) -> id canônico do handler
BUTTON_ALIASES = {
    "Quero saber mais": "interesse_direto",
    "Passei por dificuldades": "dor",
    "Não tenho interesse": "nao_interessado",
    # versões slug (se o provider mandar só o texto e você usar to_button_id)
    "quero_saber_mais": "interesse_direto",
    "passei_por_dificuldades": "dor",
    "nao_tenho_interesse": "nao_interessado",
}

def resolve_action(reply_id: str):
    if not reply_id:
        return None
    # 1) match direto
    if reply_id in BUTTON_ACTIONS:
        return BUTTON_ACTIONS[reply_id]
    # 2) alias por texto exato
    if reply_id in BUTTON_ALIASES:
        alias = BUTTON_ALIASES[reply_id]
        return BUTTON_ACTIONS.get(alias)
    # 3) tenta slug
    slug = to_button_id(reply_id)
    if slug in BUTTON_ACTIONS:
        return BUTTON_ACTIONS[slug]
    if slug in BUTTON_ALIASES:
        alias = BUTTON_ALIASES[slug]
        return BUTTON_ACTIONS.get(alias)
    return None


# Variáveis de ambiente (ou use ..env)
ACCESS_TOKEN = (os.getenv("ACCESS_TOKEN") or "").strip()
WHATSAPP_BUSINESS_ACCOUNT_ID = (os.getenv("WHATSAPP_BUSINESS_ACCOUNT_ID") or "684848121166205").strip()
PHONE_NUMBER_ID = (os.getenv("PHONE_NUMBER_ID") or "655247557670398").strip()
VERIFY_TOKEN_WEBHOOK = (os.getenv("VERIFY_TOKEN_WEBHOOK") or "2yhrqG6O4JBvT2zGXm1CWsxDadz_56XSWTU2BU6XwXcgNqnko").strip()
ATTENDANT_NUMBER = (os.getenv("ATTENDANT_NUMBER") or "").strip()
MENSAGENS_LIDAS_FILE = "mensagens_lidas.json"
# --- ENVs ---
ATTENDANT_NUMBER = os.getenv("ATTENDANT_NUMBER")              # ex: 5561999999999 (com DDI/DDD, sem +)
ATTENDANT_TEMPLATE_NAME = os.getenv("ATTENDANT_TEMPLATE_NAME", "atendimento_transfer")  # defina esse template no Meta

WINDOW_24H = 24 * 60 * 60


print("DEBUG ATTENDANT_NUMBER raw:", repr(ATTENDANT_NUMBER))

app = Flask(__name__)
app.secret_key = '6e9750a7f8050c604ba15d542bfcd5b1d2453c264b8f9770'  # 🔥 Troque por uma chave segura
CORS(app)



# 🔥 Status das buscas por usuário
status_buscas = {}

# Diretório base da aplicação (onde app.py está localizado)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define os caminhos para os arquivos de dados
DATA_DIR = os.path.join(BASE_DIR, 'data')
LEADS_FILE = os.path.join(DATA_DIR, 'leads.json')
PENDING_LEADS_FILE = os.path.join(DATA_DIR, 'pending_leads.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json') # Ajustado para usar DATA_DIR
# Diretório para exportações (planilhas geradas)
EXPORTS_DIR = os.path.join(DATA_DIR, "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)


# Caminho correto para o diretório de mensagens dentro de 'data'
MESSAGES_DIR = os.path.join(DATA_DIR, 'mensagens')

# Garante que os diretórios 'data' e 'mensagens' existam
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(MESSAGES_DIR, exist_ok=True)

# Garante que os diretórios 'data' e 'mensagens' existam
if not os.path.exists("data"):
    os.makedirs("data")
if not os.path.exists(MESSAGES_DIR):
    os.makedirs(MESSAGES_DIR)


# Defina um mapa de botões e as funções correspondentes
BUTTON_ACTIONS = {}

def register_button(id):
    """Decorator para registrar automaticamente funções de botões"""
    def wrapper(func):
        BUTTON_ACTIONS[id] = func
        return func
    return wrapper


@register_button("sim_explicar")
def trata_sim(numero):
    body = "Que ótimo! Vamos explicar melhor como funciona..."
    enviar_texto(numero, body)

@register_button("nao_explicar")
def trata_nao(numero):
    body = "Sem problemas, qualquer coisa estamos à disposição!"
    enviar_texto(numero, body)

@register_button("mais_info")
def trata_info(numero):
    body = "Aqui está o link com mais informações: https://seusite.com/info"
    enviar_texto(numero, body)

def enviar_texto(numero, body):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    msg = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "text",
        "text": {"body": body}
    }
    requests.post(url, headers=headers, json=msg)
    salvar_mensagem(numero, body, int(time.time()), remetente='sent')


# === PENDÊNCIAS DE CARDS PARA ENVIAR APÓS STATUS ===
PENDING_CARDS_FILE = os.path.join(DATA_DIR, "pending_cards.json")

def _read_pending_cards():
    if not os.path.exists(PENDING_CARDS_FILE):
        return []
    try:
        with open(PENDING_CARDS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def _write_pending_cards(items):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(PENDING_CARDS_FILE, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

def add_pending_card(template_msg_id: str, numero: str, texto: str, botoes: list):
    items = _read_pending_cards()
    items.append({
        "template_msg_id": template_msg_id,
        "numero": numero,
        "texto": texto,
        "botoes": botoes,
        "ts": int(time.time())
    })
    _write_pending_cards(items)

def pop_pending_by_msg_id(template_msg_id: str):
    items = _read_pending_cards()
    found = None
    remaining = []
    for it in items:
        if not found and it.get("template_msg_id") == template_msg_id:
            found = it
        else:
            remaining.append(it)
    _write_pending_cards(remaining)
    return found


import json as _json
def _now_ts() -> int:
    return int(time.time())

def ensure_lead(phone_e164: str, name: str | None = None):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT phone_e164 FROM leads WHERE phone_e164 = ?", (phone_e164,))
    row = cur.fetchone()
    if row is None:
        cur.execute("""
            INSERT INTO leads (phone_e164, name) VALUES (?, ?)
        """, (phone_e164, name))
    else:
        if name:
            cur.execute("UPDATE leads SET name = COALESCE(name, ?) WHERE phone_e164 = ?", (name, phone_e164))
    conn.commit(); conn.close()

def record_template_sent(phone_e164: str, template_name: str, ts: int | None = None, name: str | None = None):
    ts = ts or _now_ts()
    ensure_lead(phone_e164, name)
    conn = get_db(); cur = conn.cursor()
    # first_contacted_at só é setado se ainda não houver
    cur.execute("""
        UPDATE leads
        SET
          name = COALESCE(?, name),
          first_contacted_at = COALESCE(first_contacted_at, ?),
          last_contacted_at  = ?
        WHERE phone_e164 = ?
    """, (name, ts, ts, phone_e164))
    cur.execute("""
        INSERT INTO events (phone_e164, type, ts, meta)
        VALUES (?, 'template_sent', ?, ?)
    """, (phone_e164, ts, _json.dumps({"template": template_name})))
    conn.commit(); conn.close()

def record_initial_reply(phone_e164: str, reply_id: str, ts: int | None = None) -> bool:
    ts = ts or int(time.time())
    ensure_lead(phone_e164)
    conn = get_db(); cur = conn.cursor()

    # Marca initial_replied apenas na primeira vez; rowcount dirá se atualizou
    cur.execute("""
        UPDATE leads
           SET initial_replied = 1,
               last_reply_at   = COALESCE(last_reply_at, ?)
         WHERE phone_e164 = ?
           AND (initial_replied = 0 OR initial_replied IS NULL)
    """, (ts, phone_e164))
    first_time = (cur.rowcount > 0)

    if first_time:
        cur.execute("""
            INSERT INTO events (phone_e164, type, ts, meta)
            VALUES (?, 'initial_reply', ?, ?)
        """, (phone_e164, ts, json.dumps({"reply_id": reply_id})))

    conn.commit(); conn.close()
    return first_time


def set_outcome(phone_e164: str, outcome: str, ts: int | None = None):
    assert outcome in ("won", "lost")
    ts = ts or _now_ts()
    ensure_lead(phone_e164)
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        UPDATE leads
        SET outcome = ?, outcome_set_at = ?
        WHERE phone_e164 = ?
    """, (outcome, ts, phone_e164))
    cur.execute("""
        INSERT INTO events (phone_e164, type, ts, meta)
        VALUES (?, 'outcome_set', ?, ?)
    """, (phone_e164, ts, _json.dumps({"outcome": outcome})))
    conn.commit(); conn.close()

def summarize_metrics(start_ts: int, end_ts: int) -> dict:
    """
    start_ts inclusive, end_ts inclusive (ajuste conforme desejar)
    """
    conn = get_db(); cur = conn.cursor()
    # novos contatos (primeiro contato no período)
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM leads
        WHERE first_contacted_at IS NOT NULL
          AND first_contacted_at BETWEEN ? AND ?
    """, (start_ts, end_ts))
    novos = cur.fetchone()["c"]

    # responderam modelo inicial (dentro do mesmo conjunto acima)
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM leads
        WHERE first_contacted_at IS NOT NULL
          AND first_contacted_at BETWEEN ? AND ?
          AND initial_replied = 1
    """, (start_ts, end_ts))
    responderam = cur.fetchone()["c"]

    # não responderam (contatados no período e initial_replied = 0)
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM leads
        WHERE first_contacted_at IS NOT NULL
          AND first_contacted_at BETWEEN ? AND ?
          AND (initial_replied = 0 OR initial_replied IS NULL)
    """, (start_ts, end_ts))
    nao_resp = cur.fetchone()["c"]

    # fecharam negócio (outcome definido no período)
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM leads
        WHERE outcome = 'won'
          AND outcome_set_at BETWEEN ? AND ?
    """, (start_ts, end_ts))
    fecharam = cur.fetchone()["c"]

    # não fecharam (outcome = lost no período)
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM leads
        WHERE outcome = 'lost'
          AND outcome_set_at BETWEEN ? AND ?
    """, (start_ts, end_ts))
    nao_fecharam = cur.fetchone()["c"]

    conn.close()
    return {
        "novos_contatos_mensagens": novos,
        "responderam_modelo_inicial": responderam,
        "nao_responderam": nao_resp,
        "fecharam_negocio": fecharam,
        "nao_fecharam": nao_fecharam,
    }


def whatsapp_send_text(to: str, body: str):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": normalize_phone_number(to),
        "type": "text",
        "text": {"body": body}
    }
    r = requests.post(url, headers=headers, json=payload, timeout=60)
    try:
        j = r.json()
    except Exception:
        j = {}
    if not r.ok:
        print("ERROR send text:", r.status_code, j)
        return None
    return j


def whatsapp_send_template(to: str, template_name: str, body_params: list[str]):
    """
    Envia TEMPLATE (passa no 24h rule). body_params = ["Nome", "Telefone", "Opção"]
    Ajuste a ordem conforme seu template aprovado.
    """
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    components = [{
        "type": "body",
        "parameters": [{"type": "text", "text": str(p)[:1024]} for p in body_params]
    }]
    payload = {
        "messaging_product": "whatsapp",
        "to": normalize_phone_number(to),
        "type": "template",
        "template": {
            "name": template_name,
            "language": {"code": "pt_BR"},
            "components": components
        }
    }
    r = requests.post(url, headers=headers, json=payload, timeout=60)
    try:
        j = r.json()
    except Exception:
        j = {}
    if not r.ok:
        print("ERROR send template:", r.status_code, j)
        return None
    return j


def _last_incoming_ts(phone_e164: str) -> int:
    """
    Lê o histórico salvo em data/mensagens/<e164>.json e retorna o último timestamp
    de mensagem recebida ('received') desse contato. Se não houver, retorna 0.
    """
    try:
        num = normalize_phone_number(phone_e164)
        path = os.path.join(MESSAGES_DIR, f"{num}.json")
        if not os.path.exists(path):
            return 0
        with open(path, "r", encoding="utf-8") as f:
            msgs = json.load(f)
        last_ts = 0
        for m in msgs or []:
            if (m.get("remetente") == "received"):
                ts = int(m.get("timestamp") or m.get("ts") or 0)
                if ts > last_ts:
                    last_ts = ts
        return last_ts
    except Exception as e:
        print("WARN _last_incoming_ts:", e)
        return 0


def _resolve_lead_name_from_db_or_files(numero_e164: str, fallback: str = "Cliente") -> str:
    """
    Prioriza nome do leads.json; se não houver, tenta pending_leads.json; senão usa fallback.
    """
    try:
        # leads.json
        leads = load_leads()
        for lead in leads:
            if normalize_phone_number(lead.get("telefone")) == normalize_phone_number(numero_e164):
                n = (lead.get("nome") or lead.get("nome_lead") or "").strip()
                if n:
                    return n
                break
    except Exception as e:
        print("WARN resolve name leads.json:", e)

    # pending_leads.json
    try:
        if os.path.exists(PENDING_LEADS_FILE):
            with open(PENDING_LEADS_FILE, "r", encoding="utf-8") as f:
                pendings = json.load(f)
            if isinstance(pendings, dict):
                pendings = [pendings]
            for item in (pendings or []):
                tel = normalize_phone_number(item.get("telefone") or item.get("phone") or "")
                if tel == normalize_phone_number(numero_e164):
                    for k in ("nome", "nome_lead", "name", "contato", "responsavel"):
                        v = (item.get(k) or "").strip()
                        if v:
                            return v
                    break
    except Exception as e:
        print("WARN resolve name pending_leads.json:", e)

    return fallback


# === DB (SQLite) =============================================================
import sqlite3
DB_PATH = os.path.join(DATA_DIR, "app.db")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    # Tabela de leads: 1 linha por telefone
    cur.execute("""
    CREATE TABLE IF NOT EXISTS leads (
        phone_e164 TEXT PRIMARY KEY,
        name TEXT,
        first_contacted_at INTEGER,  -- primeira vez que enviamos template
        last_contacted_at  INTEGER,  -- última vez que enviamos template
        initial_replied   INTEGER DEFAULT 0, -- 0/1 respondeu ao modelo inicial
        last_reply_at     INTEGER,          -- quando respondeu pela 1ª vez
        outcome           TEXT CHECK (outcome IN ('won','lost') OR outcome IS NULL),
        outcome_set_at    INTEGER
    );
    """)
    # Tabela de eventos (auditoria)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        phone_e164 TEXT,
        type TEXT,            -- template_sent | initial_reply | outcome_set
        ts INTEGER,
        meta TEXT
    );
    """)
    # Índices úteis
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_ts ON events(ts);")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_leads_first ON leads(first_contacted_at);")
    conn.commit()
    conn.close()

init_db()


# --- Funções de Gerenciamento de Leads Globais ---

def load_leads():
    """Carrega leads do arquivo JSON."""
    if not os.path.exists(LEADS_FILE):
        return []
    with open(LEADS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_leads(leads):
    """Salva leads no arquivo JSON."""
    with open(LEADS_FILE, 'w', encoding='utf-8') as f:
        json.dump(leads, f, indent=2, ensure_ascii=False)

# --- Funções de Gerenciamento de Leads Pendentes ---

def load_pending_leads():
    """Carrega leads pendentes do arquivo JSON."""
    if not os.path.exists(PENDING_LEADS_FILE):
        return []
    with open(PENDING_LEADS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_pending_leads(pending_leads):
    """Salva leads pendentes no arquivo JSON."""
    with open(PENDING_LEADS_FILE, 'w', encoding='utf-8') as f:
        json.dump(pending_leads, f, indent=2, ensure_ascii=False)

def _first_nonempty(d: dict, keys: list[str]) -> Optional[str]:
    for k in keys:
        v = (d.get(k) or "").strip()
        if v:
            return v
    return None

def _load_json_list(path: str) -> list:
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            return [data]
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"DEBUG _load_json_list({path}): {e}")
    return []

def get_name_from_storage(phone_e164: str) -> Optional[str]:
    """Busca nome do lead primeiro no pending_leads.json, depois no leads.json."""
    alvo = normalize_phone_number(phone_e164)

    # 1) pending_leads.json (onde você tem 'nome')
    pendings = _load_json_list(PENDING_LEADS_FILE)
    for item in pendings:
        tel = normalize_phone_number(item.get("telefone") or item.get("phone") or "")
        if tel == alvo:
            nome = _first_nonempty(item, ["nome", "nome_lead", "name", "contato", "responsavel"])
            if nome:
                return nome

    # 2) leads.json (caso você venha a salvar nome lá depois)
    leads = load_leads()  # você já tem essa função
    for item in leads:
        tel = normalize_phone_number(item.get("telefone") or item.get("phone") or "")
        if tel == alvo:
            nome = _first_nonempty(item, ["nome", "nome_lead", "name", "contato", "responsavel"])
            if nome:
                return nome

    return None

# <<< FUNÇÃO MOVIDA E TORNADA GLOBAL >>>
def remove_pending_lead_by_phone(phone_number_to_remove):
    normalized_phone = normalize_phone_number(phone_number_to_remove)
    if not normalized_phone:
        print(f"ERROR: Não foi possível normalizar o número '{phone_number_to_remove}' para remoção de lead pendente.")
        return False

    print(f"DEBUG_REMOVE_PENDING: Tentando remover lead pendente com número normalizado: '{normalized_phone}'")
    pending_leads = load_pending_leads()
    initial_count = len(pending_leads)

    # Filtra os leads, mantendo apenas aqueles cujo telefone não é o que queremos remover
    # Garante que a comparação também seja com o número normalizado do lead salvo
    new_pending_leads = [
        lead for lead in pending_leads
        if normalize_phone_number(lead.get('telefone')) != normalized_phone
    ]

    if len(new_pending_leads) < initial_count:
        save_pending_leads(new_pending_leads)
        print(f"DEBUG_REMOVE_PENDING: Lead '{normalized_phone}' removido de pending_leads.json. Novos leads: {len(new_pending_leads)}")
        return True
    else:
        print(f"DEBUG_REMOVE_PENDING: Lead '{normalized_phone}' NÃO encontrado em pending_leads.json para remoção.")
        return False
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

def add_new_leads_to_system(new_leads):
    """
    Adiciona novos leads ao sistema (tanto leads gerais quanto pendentes),
    evitando duplicatas.
    """
    existing_leads = load_leads()
    existing_pending_leads = load_pending_leads()
    added_to_general_count = 0
    added_to_pending_count = 0

    # Identificadores de leads já existentes (geral e pendentes)
    existing_identifiers = set()
    for lead_list in [existing_leads, existing_pending_leads]:
        for lead in lead_list:
            if lead.get('site') and lead['site'] != "Site não encontrado":
                existing_identifiers.add(lead['site'])
            else:
                # Usar nome e endereço para identificar se não há site
                existing_identifiers.add((lead.get('nome'), lead.get('endereco')))

    leads_to_add_to_general = []
    leads_to_add_to_pending = []

    for lead in new_leads:
        identifier = None
        if lead.get('site') and lead['site'] != "Site não encontrado":
            identifier = lead['site']
        else:
            identifier = (lead.get('nome'), lead.get('endereco'))

        # Apenas adicione se o identificador não for None e não existir
        if identifier and identifier not in existing_identifiers:
            # ===> AQUI ESTÁ A LINHA ADICIONADA <===
            # Garante que todo novo lead adicionado já tenha o status inicial
            lead['status_contato'] = 'pendente'
            # ====================================

            leads_to_add_to_general.append(lead)
            # Adiciona à fila de pendentes se tiver telefone
            if lead.get('telefone') and lead['telefone'] != "Telefone não encontrado":
                leads_to_add_to_pending.append(lead)
            existing_identifiers.add(identifier)  # Adiciona o novo identificador ao conjunto
            added_to_general_count += 1

    if leads_to_add_to_general:
        existing_leads.extend(leads_to_add_to_general)
        save_leads(existing_leads)

    if leads_to_add_to_pending:
        existing_pending_leads.extend(leads_to_add_to_pending)
        save_pending_leads(existing_pending_leads)
        added_to_pending_count = len(leads_to_add_to_pending)  # Conta os que foram adicionados à fila

    return added_to_general_count, added_to_pending_count


def migrate_leads_status():
    """Adiciona o campo 'status_contato' e normaliza telefones de leads existentes."""
    leads = load_leads()
    pending_leads = load_pending_leads()

    updated_leads = False

    # Migra leads gerais
    for lead in leads:
        if 'status_contato' not in lead:
            lead['status_contato'] = 'novo'
            updated_leads = True
        # ===> Normaliza o telefone ao migrar <===
        if lead.get('telefone'):
            original_phone = lead['telefone']
            normalized_phone = normalize_phone_number(original_phone)
            if original_phone != normalized_phone: # Se houve alteração
                lead['telefone'] = normalized_phone
                updated_leads = True
                print(f"DEBUG: Telefone de lead {original_phone} normalizado para {normalized_phone}.")
        # ========================================

    # Migra leads pendentes
    for lead in pending_leads:
        if 'status_contato' not in lead:
            lead['status_contato'] = 'pendente'
            updated_leads = True
        # ===> Normaliza o telefone ao migrar <===
        if lead.get('telefone'):
            original_phone = lead['telefone']
            normalized_phone = normalize_phone_number(original_phone)
            if original_phone != normalized_phone:
                lead['telefone'] = normalized_phone
                updated_leads = True
                print(f"DEBUG: Telefone de lead pendente {original_phone} normalizado para {normalized_phone}.")
        # ========================================

    if updated_leads:
        save_leads(leads)
        save_pending_leads(pending_leads)
        print("DEBUG: Leads migrados e telefones normalizados para a nova estrutura.")

# === Funções auxiliares existentes ===

def salvar_mensagem(numero_original, texto, timestamp, remetente="received"):
    print(f"DEBUG_SAVE: Início de salvar_mensagem para número original: '{numero_original}' (remetente: '{remetente}')")

    numero = normalize_phone_number(numero_original) # Normaliza o número

    if not numero:
        print(f"ERROR_SAVE: Não foi possível normalizar o número para salvar mensagem: '{numero_original}'. Abortando.")
        return False # Indica que falhou

    # Constrói o caminho completo para o arquivo JSON do chat
    path = os.path.join(MESSAGES_DIR, f"{numero}.json")
    print(f"DEBUG_SAVE: Caminho do arquivo de chat esperado: '{path}'")

    mensagens = []

    # Tenta ler o arquivo JSON existente
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                mensagens = json.load(f)
            print(f"DEBUG_SAVE: Arquivo de chat existente lido com sucesso: '{path}'. Contém {len(mensagens)} mensagens.")
        except json.JSONDecodeError as e:
            print(f"ERROR_SAVE: Arquivo JSON de chat corrompido para '{numero}'. Erro: {e}. Criando um novo arquivo.")
            mensagens = [] # Se corrompido, começa com uma lista vazia
        except Exception as e:
            print(f"ERROR_SAVE: Erro inesperado ao ler arquivo de chat '{path}': {e}. Criando um novo arquivo.")
            mensagens = [] # Para outros erros de leitura

    # Adiciona a nova mensagem
    nova_mensagem = {
        "sender": remetente,
        "text": texto,
        "timestamp": timestamp
    }
    mensagens.append(nova_mensagem)
    print(f"DEBUG_SAVE: Nova mensagem adicionada à lista. Total de mensagens: {len(mensagens)}.")

    # Tenta salvar a lista de mensagens atualizada no arquivo JSON
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(mensagens, f, indent=2, ensure_ascii=False)
        print(f"DEBUG_SAVE: Mensagem de '{remetente}' para '{numero}' salva COM SUCESSO em '{path}'.")
        return True # Indica que salvou com sucesso
    except Exception as e:
        print(f"ERROR_SAVE: ERRO CRÍTICO ao salvar mensagem para '{numero}' em '{path}': {e}")
        return False # Indica que falhou


def load_users():
    """Carrega usuários do arquivo JSON."""
    users_file = USERS_FILE
    if not os.path.exists(users_file):
        # Cria um arquivo de usuários padrão se não existir
        with open(users_file, 'w', encoding='utf-8') as f:
            json.dump({"admin": {"password": "admin"}}, f, indent=2)
    with open(users_file, 'r', encoding='utf-8') as f:
        return json.load(f)

import re

def normalize_phone_number(phone_number):
    """
    Normaliza números de telefone brasileiros para o formato E.164 (+55DD9XXXXXXXX).
    Remove caracteres não numéricos.
    Garante que o '9' seja adicionado para números de celular brasileiros
    que deveriam tê-lo e que o DDI/DDD estejam presentes.
    """
    if not phone_number:
        return None

    # 1. Remove todos os caracteres não numéricos
    digits_only = re.sub(r'\D', '', phone_number)

    # 2. Garante que começa com '55' (DDI Brasil)
    if not digits_only.startswith('55'):
        # Se tem 10 ou 11 dígitos, assume que é DDD + Número e adiciona '55'
        if len(digits_only) == 10 or len(digits_only) == 11:
            digits_only = '55' + digits_only
        else:
            # Se não tem '55' e não parece um número brasileiro típico,
            # tenta retornar o que tem com '+' ou None.
            return f"+{digits_only}" if digits_only else None

    # 3. Lógica para o 9º dígito em celulares
    if len(digits_only) == 12 and digits_only.startswith('55') and len(digits_only[4:]) == 8:
        # Insere '9' após o DDD
        digits_only = digits_only[:4] + '9' + digits_only[4:]

    # Verificação de inconsistências (opcional, sem logs agora)
    if len(digits_only) < 10 or len(digits_only) > 13:
        return f"+{digits_only}" if digits_only else None

    final_number = f"+{digits_only}"
    return final_number



def check_login(username, password):
    """Verifica as credenciais de login."""
    users = load_users()
    user = users.get(username)
    if user and user['password'] == password:
        return True
    return False


# 🔥 Função que executa a busca em segundo plano
def executar_busca(username, termo, limite=50):
    try:
        status_buscas[username]["mensagem"] = "Iniciando busca..."
        status_buscas[username]["progresso"] = 5

        status_buscas[username]["mensagem"] = f"Buscando no Google Maps por: '{termo}'..."
        status_buscas[username]["progresso"] = 10

        resultado_da_busca_selenium = buscar_dados_cards_maps(
            termo=termo,
            limite=limite,
            username=username,
            status_buscas=status_buscas
        )

        status_buscas[username]["mensagem"] = "Verificando e salvando novos leads..."
        status_buscas[username]["progresso"] = 80

        added_to_general, added_to_pending = add_new_leads_to_system(resultado_da_busca_selenium)

        status_buscas[username][
            "mensagem"] = f"Busca finalizada! Total de leads encontrados: {len(resultado_da_busca_selenium)}. Novos leads gerais salvos: {added_to_general}. Novos leads adicionados à fila de contato: {added_to_pending}."
        status_buscas[username]["progresso"] = 100
        status_buscas[username]["status"] = "concluido"
        status_buscas[username]["resultado"] = resultado_da_busca_selenium

    except Exception as e:
        status_buscas[username]["mensagem"] = f"❌ Erro na busca: {str(e)}"
        status_buscas[username]["status"] = "erro"
        status_buscas[username]["progresso"] = 0


def send_whatsapp_interactive_buttons(numero, texto, botoes, context_message_id=None):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": numero,
        "type": "interactive",
        "interactive": {
            "type": "button",
            "body": {"text": texto},
            "action": {
                "buttons": [
                    {"type": "reply", "reply": {"id": b["id"], "title": b["title"]}}
                    for b in botoes
                ]
            }
        }
    }
    if context_message_id:
        payload["context"] = {"message_id": context_message_id}

    r = requests.post(url, headers=headers, json=payload, timeout=30)
    print("DEBUG send_whatsapp_interactive_buttons:", r.status_code, r.text)
    return r

def get_lead_name_by_phone(numero_e164: str) -> str | None:
    """Tenta achar o nome do lead em leads.json pelo telefone."""
    try:
        alvo = normalize_phone_number(numero_e164)
        leads = load_leads()
        for lead in leads:
            tel = normalize_phone_number(lead.get("telefone") or lead.get("phone") or lead.get("whatsapp") or "")
            if tel == alvo:
                for k in ("nome", "nome_lead", "name", "contato", "responsavel"):
                    v = (lead.get(k) or "").strip()
                    if v:
                        return v
        return None
    except Exception as e:
        print(f"DEBUG get_lead_name_by_phone: {e}")
        return None


import re

def send_whatsapp_text(to, body):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": normalize_phone_number(to),
        "type": "text",
        "text": {"body": body}
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, {"raw": r.text}

def send_whatsapp_template(to, template_name, params_body, lang="pt_BR"):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    components = []
    if params_body:
        components.append({
            "type": "body",
            "parameters": [{"type": "text", "text": str(p)[:1000]} for p in params_body]
        })
    payload = {
        "messaging_product": "whatsapp",
        "to": normalize_phone_number(to),
        "type": "template",
        "template": {
            "name": template_name,
            "language": {"code": lang},
            "components": components or None
        }
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, {"raw": r.text}



def notify_attendant_transfer(numero_lead, *args, **kwargs):
    """
    Envia TEMPLATE OFICIAL para o atendente com {{nome}}, {{numero}}, {{opcao}}
    e opcionalmente um texto de apoio na sequência.

    Aceita kwargs:
      - nome_lead
      - opcao | selection_title | selection_key
      - support_text
    """
    try:
        to_attendant  = os.getenv("ATTENDANT_NUMBER")
        template_name = os.getenv("ATTENDANT_TEMPLATE_NAME", "atendimento_transfer")
        lang_code     = os.getenv("ATTENDANT_TEMPLATE_LANG", "pt_BR")

        if not to_attendant:
            print("ATTENTION: defina ATTENDANT_NUMBER no .env")
            return False

        nome_lead   = (kwargs.get("nome_lead") or "Cliente").strip()
        opcao_str   = (kwargs.get("opcao") or kwargs.get("selection_title") or kwargs.get("selection_key") or "").strip()
        support_txt = (kwargs.get("support_text") or "").strip()

        num_e164 = normalize_phone_number(numero_lead)
        if not num_e164:
            print(f"notify_attendant_transfer: número inválido do lead: {numero_lead}")
            return False

        url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
        headers = {
            "Authorization": f"Bearer {ACCESS_TOKEN}",
            "Content-Type": "application/json",
        }

        # ---------- 1) Tenta com NAMED PARAMETERS (para {{nome}}, {{numero}}, {{opcao}}) ----------
        payload_named = {
            "messaging_product": "whatsapp",
            "to": to_attendant,
            "type": "template",
            "template": {
                "name": template_name,
                "language": {"code": lang_code},
                "components": [
                    {
                        "type": "body",
                        "parameters": [
                            {"type": "text", "text": nome_lead,      "parameter_name": "nome"},
                            {"type": "text", "text": num_e164,       "parameter_name": "numero"},
                            {"type": "text", "text": opcao_str or "-", "parameter_name": "opcao"},
                        ],
                    }
                ],
            },
        }

        r1 = requests.post(url, headers=headers, json=payload_named, timeout=30)
        try:
            j1 = r1.json()
        except Exception:
            j1 = {}

        # Se deu certo com named, segue
        if r1.ok:
            print("DEBUG notify_attendant_transfer: template (named) enviado", j1)
        else:
            # Se o erro indicar problema com 'parameter_name', tentamos POSICIONAL
            details = (j1.get("error", {}).get("error_data", {}) or {}).get("details", "")
            msg     = (j1.get("error", {}) or {}).get("message", "")
            if "parameter name" in (details.lower() + msg.lower()):
                # ---------- 2) Fallback POSICIONAL (para {{1}} {{2}} {{3}}) ----------
                payload_pos = {
                    "messaging_product": "whatsapp",
                    "to": to_attendant,
                    "type": "template",
                    "template": {
                        "name": template_name,
                        "language": {"code": lang_code},
                        "components": [
                            {
                                "type": "body",
                                "parameters": [
                                    {"type": "text", "text": nome_lead},
                                    {"type": "text", "text": num_e164},
                                    {"type": "text", "text": opcao_str or "-"},
                                ],
                            }
                        ],
                    },
                }
                r1b = requests.post(url, headers=headers, json=payload_pos, timeout=30)
                try:
                    j1b = r1b.json()
                except Exception:
                    j1b = {}
                if not r1b.ok:
                    print("ERROR notify_attendant_transfer/template (posicional):", r1b.status_code, j1b or r1b.text)
                    return False
                print("DEBUG notify_attendant_transfer: template (posicional) enviado", j1b)
            else:
                print("ERROR notify_attendant_transfer/template (named):", r1.status_code, j1 or r1.text)
                return False

        # ---------- 3) Texto de apoio (opcional) ----------
        if support_txt:
            payload_txt = {
                "messaging_product": "whatsapp",
                "to": to_attendant,
                "type": "text",
                "text": {"body": support_txt[:1024]},
            }
            r2 = requests.post(url, headers=headers, json=payload_txt, timeout=30)
            try:
                j2 = r2.json()
            except Exception:
                j2 = {}
            if r2.ok:
                print("DEBUG notify_attendant_transfer: texto enviado ao atendente", j2)
                try:
                    salvar_mensagem(to_attendant, f"[auto] {support_txt}", int(time.time()), remetente='sent')
                except Exception:
                    pass
            else:
                print("ERROR notify_attendant_transfer/text:", r2.status_code, j2 or r2.text)

        return True

    except Exception as e:
        print("notify_attendant_transfer EXCEPTION:", e)
        return False


def _safe_truncate(txt: str, limit: int = WHATSAPP_TEXT_LIMIT) -> str:
    return txt if len(txt) <= limit else (txt[: max(0, limit - 1)] + "…")

def _history_lines_from_file(numero_e164: str, max_items: int | None = None) -> list[str]:
    """
    Lê data/mensagens/<numero>.json e devolve linhas:
    - [dd/mm HH:MM] Origem: texto
    Usa seu formato atual: [{"sender": "received"|"sent", "text": "...", "timestamp": ...}, ...]
    """
    try:
        path = os.path.join(MESSAGES_DIR, f"{normalize_phone_number(numero_e164)}.json")
        if not os.path.exists(path):
            return []
        with open(path, "r", encoding="utf-8") as f:
            mensagens = json.load(f)
        if not isinstance(mensagens, list):
            return []

        if max_items:
            mensagens = mensagens[-max_items:]

        linhas = []
        for m in mensagens:
            ts = int(m.get("timestamp", 0) or 0)
            try:
                quando = datetime.fromtimestamp(ts).strftime("%d/%m %H:%M")
            except Exception:
                quando = "?"
            sender = (m.get("sender") or "").lower()
            origem = "Cliente" if sender == "received" else ("Bot" if sender == "sent" else (sender or "?"))
            texto = (m.get("text") or "").strip().replace("\r", "")
            if len(texto) > 700:
                texto = texto[:700] + "…"
            linhas.append(f"- [{quando}] {origem}: {texto}")
        return linhas
    except Exception as e:
        print("WARN _history_lines_from_file:", e)
        return []

def _chunk_and_send_attendant(header: str, lines: list[str], attendant_to: str):
    """
    Envia em múltiplas mensagens se exceder limite. Primeira contém header + início do histórico.
    """
    MAX = 3500  # margem para não bater 4096
    partes = []
    cur = header.strip()

    if lines:
        cur += "\n\n— *Histórico* —\n"

    for ln in lines or []:
        add = ("\n" if cur else "") + ln
        if len(cur) + len(add) > MAX:
            partes.append(cur)
            cur = ln
        else:
            cur += add

    if cur.strip():
        partes.append(cur.strip())

    total = len(partes) or 1
    for i, p in enumerate(partes, 1):
        body = p if total == 1 else f"{p}\n\n(parte {i}/{total})"
        whatsapp_send_text(attendant_to, _safe_truncate(body))

def notify_attendant_history(
    numero_lead: str,
    nome_lead: str,
    selection_title: str,
    support_text: str = "",
    history: str = "last",   # "full" | "last" | "none"
    last_n: int = 20
):
    """
    Envia uma MENSAGEM DE TEXTO (sem template) para o atendente com:
    cabeçalho + link wa.me + (opcional) histórico (com quebra em partes).
    Deve ser chamada DEPOIS do template oficial (notify_attendant_transfer).
    """
    try:
        attendant = os.getenv("ATTENDANT_NUMBER") or ATTENDANT_NUMBER
        if not attendant:
            print("ATTENTION: defina ATTENDANT_NUMBER no .env")
            return False

        lead_e164  = normalize_phone_number(numero_lead)
        lead_digits = re.sub(r"\D", "", lead_e164)
        wa_url = f"https://wa.me/{lead_digits}"

        header = (
            "🔔 *Novo lead para atendimento*\n"
            f"*Nome:* {nome_lead or 'Cliente'}\n"
            f"*Número:* {lead_e164}\n"
            f"*Opção:* {selection_title or '-'}\n\n"
            "*Mensagem enviada ao cliente:*\n"
            f"{(support_text or '-').strip()}\n\n"
            f"Abrir chat: {wa_url}"
        )

        if history == "none":
            whatsapp_send_text(attendant, _safe_truncate(header))
            return True

        lines = _history_lines_from_file(lead_e164, None if history == "full" else last_n)
        _chunk_and_send_attendant(header, lines, attendant)
        return True
    except Exception as e:
        print("ERROR notify_attendant_history:", e)
        return False

def notify_attendant_forward_message(numero_lead: str, text: str, nome_lead: str = "", attach_last: int = 6):
    """
    Encaminha uma nova mensagem do cliente para o atendente com histórico recente junto.
    """
    try:
        attendant = os.getenv("ATTENDANT_NUMBER") or ATTENDANT_NUMBER
        if not attendant:
            print("ATTENTION: defina ATTENDANT_NUMBER no .env")
            return False

        lead_e164  = normalize_phone_number(numero_lead)
        lead_digits = re.sub(r"\D", "", lead_e164)
        wa_url = f"https://wa.me/{lead_digits}"

        header = (
            f"📨 *Nova mensagem de {nome_lead or 'Cliente'}*\n"
            f"{lead_e164}\n\n"
            f"{(text or '-').strip()}\n\n"
            f"Abrir chat: {wa_url}"
        )
        lines = _history_lines_from_file(lead_e164, attach_last)
        _chunk_and_send_attendant(header, lines, attendant)
        return True
    except Exception as e:
        print("ERROR notify_attendant_forward_message:", e)
        return False
# ============================================================================ #



# === ROTAS DE LOGIN ===

@app.route('/')
def index():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        username = request.form.get('username') or data.get('username')
        password = request.form.get('password') or data.get('password')

        if check_login(username, password):
            session['username'] = username

            # Se veio JSON (AJAX), responde JSON.
            # Se veio formulário (html form), faz redirect.
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"status": "success", "redirect": url_for('dashboard')}), 200
            return redirect(url_for('dashboard'))

        # Para erro, se for AJAX devolve JSON com 401; senão renderiza login com erro.
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"status": "error", "message": "Usuário ou senha inválidos"}), 401
        return render_template('login.html', erro="Usuário ou senha inválidos")

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))


# === DASHBOARD ===

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html', username=session['username'])

@app.route('/conversas')
def conversas():
    # Isso instrui o Flask a procurar 'conversas.html' na pasta 'templates'
    return render_template('conversas.html')
# === Middleware de proteção ===

def login_required(func):
    """Decorador para proteger rotas que exigem login."""
    from functools import wraps # Importa aqui para garantir que está disponível
    @wraps(func)
    def wrapper(*args, **kwargs):
        if 'username' not in session:
            # Para APIs, retornar JSON; para páginas, redirecionar
            if request.path.startswith('/api/'):
                return jsonify({"error": "Não autorizado"}), 401
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    return wrapper


# === APIs ===

@app.route('/api/iniciar-busca', methods=['POST'])
@login_required
def iniciar_busca_api():
    data = request.json
    termo = data.get("termo")
    limite = data.get("limite", 5)

    if not termo:
        return jsonify({"error": "Termo não informado"}), 400

    username = session['username']

    if username in status_buscas and status_buscas[username]["status"] == "buscando":
        return jsonify({"mensagem": "Já existe uma busca em andamento."}), 409 # Conflict

    status_buscas[username] = {
        "status": "buscando",
        "mensagem": "Iniciando busca...",
        "progresso": 0,
        "parciais": [],
        "resultado": []
    }

    thread = threading.Thread(target=executar_busca, args=(username, termo, limite))
    thread.start()

    return jsonify({"mensagem": "Busca iniciada com sucesso"})


@app.route('/api/status-busca', methods=['GET'])
@login_required
def status_busca_api():
    username = session['username']
    status = status_buscas.get(username, {
        "status": "parado",
        "mensagem": "Nenhuma busca em andamento",
        "progresso": 0,
        "parciais": [],
        "resultado": []
    })
    return jsonify(status)


@app.route('/api/resetar-busca', methods=['POST'])
@login_required
def resetar_busca_api():
    username = session['username']
    status_buscas.pop(username, None)
    return jsonify({"mensagem": "Busca resetada com sucesso"})


# 🔍 APIs auxiliares opcionais

@app.route('/api/buscar-links-sites', methods=['POST'])
@login_required
def buscar_links_sites_api():
    data = request.json
    termo = data.get("termo")

    if not termo:
        return jsonify({"error": "Termo não informado"}), 400

    links = buscar_links_site_maps(termo)
    return jsonify(links)


@app.route('/api/scrapear-contato', methods=['POST'])
@login_required
def scrapear_contato_api():
    data = request.json
    url = data.get("url")

    if not url:
        return jsonify({"error": "URL não informada"}), 400

    contatos = extrair_contatos(url)
    return jsonify(contatos)


@app.route('/api/chats', methods=['GET'])
@login_required
def listar_chats_api():
    if not os.path.exists(MESSAGES_DIR):
        return jsonify([])

    arquivos = [
        f.replace('.json', '')
        for f in os.listdir(MESSAGES_DIR)
        if f.endswith('.json')
    ]
    return jsonify(arquivos)


@app.route("/leads")
def leads():
    """
    Renderiza a página de leads.
    """
    # Você pode passar variáveis para o template aqui, se necessário.
    # Ex: leads_data = obter_dados_dos_leads_do_banco_de_dados()
    # return render_template('leads.html', leads=leads_data)
    return render_template('leads.html')


@app.route('/api/mensagens', methods=['GET'])
@login_required
def mensagens_api():
    numero_original = request.args.get('numero')
    numero = normalize_phone_number(numero_original)

    if not numero:
        return jsonify({"erro": "Número não informado ou inválido após normalização"}), 400

    path = os.path.join(MESSAGES_DIR, f"{numero}.json")
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                mensagens = json.load(f)

            # Carrega último timestamp lido para este número
            mensagens_lidas = carregar_mensagens_lidas()
            ultimo_lido = mensagens_lidas.get(numero, 0)

            mensagens_com_status = []
            maior_timestamp = ultimo_lido

            for msg in mensagens:
                msg_copy = msg.copy()
                ts = int(msg.get('timestamp', 0))
                if msg.get('sender') == 'received' and ts > ultimo_lido:
                    msg_copy['nova'] = True
                    # guarda o maior timestamp novo encontrado
                    if ts > maior_timestamp:
                        maior_timestamp = ts
                else:
                    msg_copy['nova'] = False
                mensagens_com_status.append(msg_copy)

            # 🔑 Atualiza o ultimo_lido só se encontrou novas
            if maior_timestamp > ultimo_lido:
                mensagens_lidas[numero] = maior_timestamp
                salvar_mensagens_lidas(mensagens_lidas)

            return jsonify(mensagens_com_status)

        except json.JSONDecodeError as e:
            return jsonify({"erro": f"Arquivo JSON corrompido: {e}"}), 500

    return jsonify([])


# ** Este endpoint NÃO é para enviar mensagens do frontend, é um webhook para receber da Meta **
# Foi renomeado de 'enviar_mensagem' para 'enviar_mensagem_whatsapp_api' para clareza,
# mas o endpoint correto para mensagens DE TEXTO LIVRE do frontend é 'enviar_mensagem_personalizada'
# ou 'enviar_template'.
@app.route('/api/enviar-mensagem-padrao', methods=['POST'])
@login_required
def enviar_mensagem_padrao_api():
    """
    OBS: Este endpoint agora é 'enviar-mensagem-padrao'.
    Se o seu frontend ainda chama '/api/enviar-mensagem', ele precisa ser atualizado
    para '/api/enviar-mensagem-personalizada' ou '/api/enviar-template'.
    Esta função foi movida para que o nome '/api/enviar-mensagem' fique livre
    para um futuro uso de Webhook ou outro propósito mais claro.
    """
    data = request.json
    numero = data.get('numero')
    mensagem = data.get('mensagem') # A mensagem completa vem do frontend

    if not numero or not mensagem:
        return jsonify({"erro": "Número e mensagem são obrigatórios"}), 400

    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    payload = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "text",
        "text": {
            "body": mensagem
        }
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        response_data = response.json()

        if response.ok:
            salvar_mensagem(numero, mensagem, int(time.time()), remetente='sent')
            # remove_pending_lead_by_phone(numero) # Removido daqui, pois o chat normal não remove o lead pendente automaticamente
            return jsonify(response_data), 200
        else:
            print(f"Erro da API do WhatsApp ao enviar mensagem padrão: {response_data}")
            error_message = response_data.get('error', {}).get('message', 'Erro desconhecido da API do WhatsApp.')
            return jsonify({"erro": error_message}), response.status_code
    except Exception as e:
        print(f"Erro na requisição para a API do WhatsApp (padrão): {str(e)}")
        return jsonify({"erro": f"Falha interna ao enviar mensagem padrão: {str(e)}"}), 500


# RECEBER MENSAGENS (WEBHOOK)

@app.route('/webhook', methods=['GET', 'POST'])
def webhook_whatsapp():
    if request.method == 'GET':
        # Verificação do webhook
        mode = request.args.get('hub.mode')
        token = request.args.get('hub.verify_token')
        challenge = request.args.get('hub.challenge')

        if mode and token:
            if mode == 'subscribe' and token == VERIFY_TOKEN_WEBHOOK:
                print('WEBHOOK_VERIFIED')
                return Response(challenge, mimetype='text/plain'), 200
            else:
                return 'Forbidden', 403
        else:
            return 'Bad Request', 400

    elif request.method == 'POST':
        data = request.get_json()
        print("🔔 Nova mensagem recebida (Webhook POST):")
        pprint(data)

        # ---------------- COOLDOWNS (ajustáveis via .env) ----------------
        TRANSFER_TTL = int(os.getenv("ATTENDANT_TRANSFER_TTL", "900"))   # 15 min
        FORWARD_TTL  = int(os.getenv("ATTENDANT_FORWARD_TTL", "60"))     # 60 s

        # ------- helpers locais para consultar/registrar eventos no SQLite
        def _recent_event_within(phone_e164: str, types: tuple[str, ...], ttl_seconds: int) -> bool:
            """Retorna True se existe evento de um desses types dentro do TTL."""
            try:
                now = int(time.time())
                conn = get_db(); cur = conn.cursor()
                placeholders = ",".join("?" * len(types))
                cur.execute(
                    f"SELECT MAX(ts) AS last FROM events WHERE phone_e164 = ? AND type IN ({placeholders})",
                    (phone_e164, *types)
                )
                row = cur.fetchone()
                conn.close()
                last = row[0] if row and row[0] is not None else None
                return bool(last and (now - int(last) < ttl_seconds))
            except Exception as e:
                print("WARN _recent_event_within:", e)
                return False

        def _insert_event(phone_e164: str, etype: str, meta: dict | None = None, ts: int | None = None):
            try:
                ts = ts or int(time.time())
                conn = get_db(); cur = conn.cursor()
                cur.execute(
                    "INSERT INTO events (phone_e164, type, ts, meta) VALUES (?,?,?,?)",
                    (phone_e164, etype, ts, json.dumps(meta or {}))
                )
                conn.commit(); conn.close()
            except Exception as e:
                print("WARN _insert_event:", e)

        def _wa_link(e164: str) -> str:
            digits = re.sub(r"\D", "", e164 or "")
            return f"https://wa.me/{digits}"

        if data and 'object' in data and 'entry' in data:
            for entry in data.get('entry', []):
                for change in entry.get('changes', []):
                    if change.get('field') != 'messages':
                        continue

                    value = change.get('value', {}) or {}

                    # ====== HANDLER DE STATUS (template enviado/entregue) ======
                    statuses = value.get('statuses', []) or []
                    for st in statuses:
                        st_id = st.get('id')            # ID da mensagem enviada (o template_msg_id)
                        st_status = st.get('status')    # 'sent' | 'delivered' | 'read' | ...
                        print(f"DEBUG_STATUS: Recebido status para {st_id}: {st_status}")

                        if st_id and st_status in ('sent', 'delivered'):
                            pending = pop_pending_by_msg_id(st_id)
                            print(f"DEBUG_STATUS: Pending encontrado? {bool(pending)} para {st_id}")
                            if pending:
                                try:
                                    send_whatsapp_interactive_buttons(
                                        pending["numero"],
                                        pending["texto"],
                                        pending["botoes"],
                                        context_message_id=st_id  # encadeia no template
                                    )
                                    print(f"DEBUG_STATUS: Card enviado após status {st_status} para {st_id}")
                                except Exception as e:
                                    print(f"Falha ao enviar card pendente {st_id}: {e}")

                    # ====== HANDLER DE MENSAGENS ======
                    messages = value.get('messages', []) or []
                    contacts = value.get('contacts', []) or []

                    # Mapa de nomes por wa_id normalizado
                    contact_names = {
                        normalize_phone_number(c.get('wa_id')): c.get('profile', {}).get('name', c.get('wa_id'))
                        for c in contacts
                    }

                    # labels “bonitos” para primeiro encaminhamento sem botão
                    label_by_type = {
                        "text": "Texto livre",
                        "image": "Imagem",
                        "video": "Vídeo",
                        "audio": "Áudio",
                        "document": "Documento",
                        "location": "Localização",
                        "contacts": "Contato",
                        "reaction": "Reação",
                    }

                    for mensagem in messages:
                        numero = normalize_phone_number(mensagem.get('from'))
                        mensagem_id = mensagem.get('id')
                        timestamp = int(mensagem.get('timestamp') or 0)
                        tipo_mensagem = mensagem.get('type')
                        nome_contato = contact_names.get(numero, numero)

                        texto = None
                        reply_id = None
                        reply_title = None

                        # ========= TIPOS DE MENSAGEM =========
                        if tipo_mensagem == 'text':
                            texto = (mensagem.get('text', {}) or {}).get('body')

                            # Gatilho opcional de menu
                            gatilho = (texto or '').strip().lower()
                            wants_menu = any(k in gatilho for k in [
                                'atendente', 'falar com atendente', 'menu',
                                'opcoes', 'opções', 'ajuda', 'suporte'
                            ])
                            if wants_menu:
                                try:
                                    send_whatsapp_interactive_buttons(
                                        numero,
                                        "Deseja falar com um atendente?",
                                        [
                                            {"id": "sim_atendente", "title": "Sim"},
                                            {"id": "nao_atendente", "title": "Não"}
                                        ]
                                    )
                                except Exception as e:
                                    print(f"Falha ao enviar botões automáticos: {e}")

                        elif tipo_mensagem == 'button':
                            btn = mensagem.get('button', {}) or {}
                            reply_title = (btn.get('text') or '').strip()
                            reply_id = (btn.get('payload') or '').strip() or to_button_id(reply_title)
                            texto = f"Resposta do botão: {reply_title} (ID: {reply_id})"

                        elif tipo_mensagem == 'interactive':
                            interactive_data = mensagem.get('interactive', {}) or {}
                            itype = interactive_data.get('type')

                            if itype == 'list_reply':
                                lr = interactive_data.get('list_reply', {}) or {}
                                reply_id = lr.get('id')
                                reply_title = lr.get('title')
                                texto = f"Resposta da lista: {reply_title} (ID: {reply_id})"

                            elif itype == 'button_reply':
                                br = interactive_data.get('button_reply', {}) or {}
                                reply_id = br.get('id')
                                reply_title = br.get('title')
                                texto = f"Resposta do botão: {reply_title} (ID: {reply_id})"

                        elif tipo_mensagem == 'reaction':
                            emoji = (mensagem.get('reaction', {}) or {}).get('emoji')
                            texto = f"Reação: {emoji}"

                        elif tipo_mensagem == 'image':
                            texto = "Imagem recebida"

                        elif tipo_mensagem == 'video':
                            texto = "Vídeo recebido"

                        elif tipo_mensagem == 'audio':
                            texto = "Áudio recebido"

                        elif tipo_mensagem == 'document':
                            texto = "Documento recebido"

                        elif tipo_mensagem == 'location':
                            texto = "Localização recebida"

                        elif tipo_mensagem == 'contacts':
                            texto = "Contato(s) recebido(s)"

                        # ========= PÓS-PROCESSAMENTO (salvar + atualizar lead) =========
                        if numero and texto:
                            salvar_mensagem(numero, texto, timestamp, remetente='received')
                            print(f"DEBUG: Mensagem de {numero} salva: {texto}")

                            # Marca "respondeu ao modelo inicial"
                            first = False
                            if tipo_mensagem == 'text':
                                try:
                                    first = record_initial_reply(numero, reply_id or 'free_text', timestamp)
                                    if first:
                                        print(f"DEBUG DB: initial_reply (texto) gravado para {numero}")
                                except Exception as e:
                                    print("WARN record_initial_reply (texto):", e)
                            else:
                                if not reply_id:
                                    try:
                                        first = record_initial_reply(numero, f'free_{tipo_mensagem}', timestamp)
                                        if first:
                                            print(f"DEBUG DB: initial_reply ({tipo_mensagem}) gravado para {numero}")
                                    except Exception as e:
                                        print("WARN record_initial_reply (mídia):", e)

                            # Atualiza status para em_conversacao
                            leads = load_leads()
                            updated = False
                            for lead in leads:
                                if normalize_phone_number(lead.get('telefone')) == numero:
                                    if lead.get('status_contato') in ["pendente", "novo", "contatado", None]:
                                        lead['status_contato'] = "em_conversacao"
                                        updated = True
                                    break
                            if updated:
                                save_leads(leads)

                            # ========= ROTEAMENTO COM LIMITES =========
                            if reply_id:
                                # Clique em botão/lista -> segue fluxo (essas ações já chamam notify_attendant_transfer)
                                try:
                                    acao = resolve_action(reply_id)
                                    if acao:
                                        acao(
                                            numero,
                                            reply_title=reply_title,
                                            context_wamid=mensagem_id,
                                            nome_contato=nome_contato
                                        )
                                    else:
                                        print(f"Nenhuma ação registrada para reply_id={reply_id} (silencioso)")
                                except Exception as e:
                                    print(f"Falha ao processar reply_id '{reply_id}': {e}")
                            else:
                                # Mensagem avulsa (sem reply_id)
                                nome_para_card = get_lead_name_by_phone(numero) or nome_contato or "Cliente"
                                titulo = label_by_type.get(tipo_mensagem, "Mensagem")

                                # 1) Primeira resposta do cliente: tenta abrir card, mas respeita TTL
                                if first:
                                    if not _recent_event_within(numero, ("attendant_transfer",), TRANSFER_TTL):
                                        notify_attendant_transfer(
                                            numero_lead=numero,
                                            selection_key=f"free_{tipo_mensagem}",
                                            selection_title=titulo,
                                            support_text=texto,
                                            nome_lead=nome_para_card
                                        )
                                        _insert_event(numero, "attendant_transfer", {
                                            "source": f"free_{tipo_mensagem}",
                                            "title": titulo
                                        }, ts=timestamp)
                                    else:
                                        print("COOLDOWN: pulo template de transferência (attendant_transfer) dentro do TTL.")

                                # 2) Encaminhar atualizações subsequentes com cooldown
                                else:
                                    if ATTENDANT_NUMBER and not _recent_event_within(numero, ("attendant_forward",), FORWARD_TTL):
                                        try:
                                            forward_text = (
                                                "📝 *Atualização do lead*\n"
                                                f"• *Nome:* {nome_para_card}\n"
                                                f"• *Número:* {numero}\n"
                                                f"• *Mensagem:* {texto}\n\n"
                                                f"Abrir chat: {_wa_link(numero)}"
                                            )
                                            whatsapp_send_text(ATTENDANT_NUMBER, forward_text)
                                            _insert_event(numero, "attendant_forward", {
                                                "len": len(texto or ""),
                                                "title": titulo
                                            }, ts=timestamp)
                                        except Exception as e:
                                            print("WARN forward attendant:", e)
                                    else:
                                        print("COOLDOWN: pulo encaminhamento (attendant_forward) dentro do TTL.")
                        else:
                            print(
                                f"Alerta: Mensagem do webhook sem número ou texto. Tipo: {tipo_mensagem}. Dados: {mensagem}"
                            )

        return 'OK', 200

    return 'Method Not Allowed', 405




@app.route("/api/leads/outcome", methods=["POST"])
@login_required
def api_set_outcome():
    data = request.get_json() or {}
    numero = normalize_phone_number(data.get("numero") or "")
    outcome = (data.get("outcome") or "").strip().lower()  # 'won' ou 'lost'
    if not numero or outcome not in ("won","lost"):
        return jsonify({"erro":"Parâmetros inválidos"}), 400
    set_outcome(numero, outcome, int(time.time()))
    return jsonify({"ok": True}), 200

@app.route("/api/leads/list", methods=["GET"])
@login_required
def api_leads_list():
    start = request.args.get("start")  # "2025-08-01"
    end   = request.args.get("end")    # "2025-08-31"
    if not start or not end:
        return jsonify({"erro": "Passe start/end no formato YYYY-MM-DD"}), 400

    start_ts = _parse_date_ymd(start)
    end_ts   = _parse_date_ymd(end) + 86399

    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT phone_e164, COALESCE(name,'') AS name,
                   first_contacted_at, initial_replied, outcome, outcome_set_at
            FROM leads
            WHERE first_contacted_at IS NOT NULL
              AND first_contacted_at BETWEEN ? AND ?
            ORDER BY first_contacted_at DESC
        """, (start_ts, end_ts))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"items": rows}), 200
    except Exception as e:
        print("ERROR /api/leads/list:", e)
        return jsonify({"erro": "Falha ao listar leads"}), 500

from datetime import datetime

def _parse_date_ymd(s: str) -> int:
    # retorna timestamp (UTC) do início do dia
    dt = datetime.strptime(s, "%Y-%m-%d")
    return int(dt.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())

@app.route("/api/metrics/summary", methods=["GET"])
@login_required
def api_metrics_summary():
    start = request.args.get("start")  # "2025-08-01"
    end   = request.args.get("end")    # "2025-08-31"
    if not start or not end:
        return jsonify({"erro":"Passe start/end no formato YYYY-MM-DD"}), 400

    start_ts = _parse_date_ymd(start)
    # end_ts = fim do dia
    end_ts   = _parse_date_ymd(end) + 86399

    try:
        data = summarize_metrics(start_ts, end_ts)
        return jsonify({"range": {"start": start, "end": end}, "data": data}), 200
    except Exception as e:
        print("ERROR summary:", e)
        return jsonify({"erro": "Falha ao gerar resumo"}), 500


# --- Helpers e roteador de botões ---

BUTTON_ACTIONS = {}

@register_button("interesse_direto")
def acao_interesse_direto(numero, **kwargs):
    body = (
        "Que ótimo! 😃\n"
        "Para entender melhor como podemos ajudar, me fala um pouco sobre o seu negócio: "
        "qual é o segmento e quais os principais desafios que você enfrenta hoje com contabilidade ou finanças?\n\n"
        "Assim eu consigo te mostrar uma solução personalizada. 🚀"
    )
    enviar_texto(numero, body)

    nome = (kwargs.get("nome_contato")
            or get_lead_name_by_phone(numero)
            or "Cliente")

    # 1) template oficial para abrir a janela com o atendente
    notify_attendant_transfer(
        numero_lead=numero,
        selection_key="interesse_direto",
        selection_title="Quero saber mais",
        support_text=body,
        nome_lead=nome
    )

    # 2) texto com histórico completo (em partes, se necessário)
    notify_attendant_history(
        numero_lead=numero,
        nome_lead=nome,
        selection_title="Quero saber mais",
        support_text=body,
        history="full",   # "last" para ficar mais curto
        last_n=20
    )



@register_button("dor")
def acao_dor(numero, **kwargs):
    body = (
        "Entendi! É muito comum aparecerem esses desafios mesmo.\n\n"
        "Se puder me contar um pouco mais sobre qual foi a maior dificuldade que você teve (ou tem hoje), "
        "consigo te mostrar como a Bloco 244 pode simplificar isso pra você. 😉"
    )
    enviar_texto(numero, body)

    nome = (kwargs.get("nome_contato")
            or get_lead_name_by_phone(numero)
            or "Cliente")

    notify_attendant_transfer(
        numero_lead=numero,
        selection_key="dor",
        selection_title="Passei por dificuldades",
        support_text=body,
        nome_lead=nome
    )

    notify_attendant_history(
        numero_lead=numero,
        nome_lead=nome,
        selection_title="Passei por dificuldades",
        support_text=body,
        history="full",
        last_n=20
    )





@register_button("nao_interessado")
def acao_nao_interessado(numero, **kwargs):
    texto = (
        "Sem problema, eu entendo 👍\n"
        "Se em algum momento precisar de apoio com contabilidade ou organização financeira, "
        "a Bloco 244 estará à disposição.\n\n"
        "Posso manter seu contato para compartilhar, de vez em quando, dicas rápidas que podem ajudar "
        "na gestão do seu negócio? 📩"
    )

    try:
        send_whatsapp_interactive_buttons(
            numero,
            texto,
            [
                {"id": "opt_in_newsletter",  "title": "📬 Enviar"},
                {"id": "opt_out_newsletter", "title": "❌ Não receber"},
            ],
            context_message_id=kwargs.get("context_wamid")
        )
    except Exception as e:
        print(f"Falha ao enviar caixinha de 2 botões (opt-in/out): {e}")



@register_button("opt_in_newsletter")
def acao_opt_in_news(numero, **kwargs):
    leads = load_leads()
    for lead in leads:
        if normalize_phone_number(lead.get("telefone")) == numero:
            lead["consent_marketing"] = True
            break
    save_leads(leads)
    enviar_texto(
        numero,
        "Perfeito! ✅ Vou te enviar conteúdos curtos e úteis de vez em quando. "
        "Se quiser parar a qualquer momento, é só me avisar. 😉"
    )

@register_button("opt_out_newsletter")
def acao_opt_out_news(numero, **kwargs):
    leads = load_leads()
    for lead in leads:
        if normalize_phone_number(lead.get("telefone")) == numero:
            lead["consent_marketing"] = False
            break
    save_leads(leads)
    enviar_texto(
        numero,
        "Tudo bem! ❌ Não vou enviar conteúdos. "
        "Se mudar de ideia, é só me mandar uma mensagem. 🙂"
    )




@app.route('/api/whatsapp_templates', methods=['GET'])
@login_required
def get_whatsapp_templates():
    # Use WHATSAPP_BUSINESS_ACCOUNT_ID para buscar templates
    url = f"https://graph.facebook.com/v18.0/{WHATSAPP_BUSINESS_ACCOUNT_ID}/message_templates" #
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}"
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status() # Lança exceção para erros HTTP
        templates_data = response.json()

        approved_templates = []
        for template in templates_data.get('data', []):
            if template.get('status') == 'APPROVED' and template.get('category') in ['UTILITY', 'MARKETING']:
                approved_templates.append({
                    "name": template.get('name'),
                    "language": template.get('language'),
                    "category": template.get('category'),
                    # Se precisar de informações sobre os parâmetros do template, adicione aqui
                    # "components": template.get('components', [])
                })
        return jsonify(approved_templates), 200

    except requests.exceptions.RequestException as e:
        print(f"Erro ao buscar templates do WhatsApp: {e}")
        return jsonify({"erro": f"Erro ao buscar modelos de template: {str(e)}"}), 500
    except Exception as e:
        print(f"Erro inesperado ao processar templates: {e}")
        return jsonify({"erro": f"Erro interno: {str(e)}"}), 500


@app.route('/api/enviar-template', methods=['POST'])
@login_required
def enviar_template_api():
    data = request.json
    numero_original = data.get('numero')
    nome_lead_req   = (data.get('nome_lead') or '').strip()
    template_name   = data.get('template_name')

    if not numero_original or not template_name:
        return jsonify({"erro": "Número e nome do template são obrigatórios."}), 400

    numero = normalize_phone_number(numero_original)
    if not numero:
        return jsonify({"erro": "Número é inválido após normalização"}), 400

    # --- 1) Tenta pegar o nome do pending_leads.json se não veio no request ---
    nome_from_pending = None
    try:
        if os.path.exists(PENDING_LEADS_FILE):
            with open(PENDING_LEADS_FILE, "r", encoding="utf-8") as f:
                pendings = json.load(f)
            if isinstance(pendings, dict):
                pendings = [pendings]
            for item in (pendings or []):
                tel = normalize_phone_number(item.get("telefone") or item.get("phone") or "")
                if tel == numero:
                    # tenta campos comuns de nome
                    for k in ("nome", "nome_lead", "name", "contato", "responsavel"):
                        v = (item.get(k) or "").strip()
                        if v:
                            nome_from_pending = v
                            break
                    break
    except Exception as e:
        print(f"DEBUG carregar pending_leads para nome: {e}")

    # Prioridade: nome do request > pending_leads.json > fallback
    nome_final = nome_lead_req or nome_from_pending or "Cliente"

    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    # --- 2) Template COM {{1}} -> envia 1 parâmetro no body (nome_final) ---
    payload = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "template",
        "template": {
            "name": template_name,
            "language": {"code": "pt_BR"},
            "components": [
                {
                    "type": "body",
                    "parameters": [
                        {"type": "text", "text": nome_final}
                    ]
                }
            ]
        }
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        response_data = response.json()

        if response.ok:
            # 1) ID da mensagem do template (se precisar correlacionar em logs)
            try:
                template_msg_id = (response_data.get('messages') or [{}])[0].get('id')
            except Exception:
                template_msg_id = None

            # 2) histórico
            salvar_mensagem(
                numero,
                f"Template '{template_name}' enviado para {nome_final}.",
                int(time.time()),
                remetente='sent'
            )

            # 2.1) registra no banco  <-- NOVO
            try:
                record_template_sent(numero, template_name, int(time.time()), name=nome_final)
                print("DEBUG DB: record_template_sent OK", numero, template_name, nome_final)
            except Exception as e:
                print("WARN DB: falha ao registrar template_sent:", e)

            # 3) status do lead + **garantir persistência do NOME no leads.json**
            leads = load_leads()
            found = None
            for lead in leads:
                if normalize_phone_number(lead.get('telefone')) == numero:
                    found = lead
                    break

            updated = False
            if found:
                # atualiza status
                if found.get('status_contato') in ["pendente", "novo", None]:
                    found['status_contato'] = "contatado"
                    updated = True
                    print(f"DEBUG: Status do lead {numero} atualizado para 'contatado' (via template).")
                # grava nome se não existir
                if not (found.get('nome') or found.get('nome_lead')):
                    found['nome'] = nome_final
                    updated = True
            else:
                # cria lead mínimo se não existir
                leads.append({
                    "telefone": numero,
                    "nome": nome_final,
                    "status_contato": "contatado"
                })
                updated = True

            if updated:
                save_leads(leads)

            # 4) remover do pending (se existir)
            remove_pending_lead_by_phone(numero)

            # 5) Não agendamos mais card aqui — o template tem seus próprios botões
            return jsonify(response_data), 200

        else:
            print(f"Erro da API do WhatsApp ao enviar template: {response_data}")
            error_message = response_data.get('error', {}).get('message', 'Erro desconhecido da API do WhatsApp.')
            return jsonify({"erro": error_message}), response.status_code

    except Exception as e:
        print(f"Erro na requisição para a API do WhatsApp (template): {str(e)}")
        return jsonify({"erro": f"Falha interna ao enviar template: {str(e)}"}), 500




# Página do painel de métricas
@app.route("/metrics")
@login_required
def metrics():                       # << endpoint = "metrics"
    # passe o username como já faz nas outras telas
    return render_template("metrics.html", username=session.get("username", "Usuário"))



## **NOVO ENDPOINT PARA MENSAGENS PERSONALIZADAS (TEXTO LIVRE)**

@app.route('/api/enviar-mensagem-personalizada', methods=['POST'])
@login_required
def enviar_mensagem_personalizada_api():
    data = request.json
    numero_cliente_original = data.get('numero') # Recebe o número do frontend
    mensagem_texto = data.get('mensagem')

    # <<<< IMPORTANTE: NORMALIZAR O NÚMERO ANTES DE USÁ-LO >>>>
    numero_cliente_normalizado = normalize_phone_number(numero_cliente_original)

    if not numero_cliente_normalizado:
        return jsonify({"erro": "Número do cliente inválido ou não normalizável."}), 400

    if not mensagem_texto:
        return jsonify({"erro": "Mensagem vazia."}), 400

    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": numero_cliente_normalizado,
        "type": "text",
        "text": {
            "body": mensagem_texto
        }
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        response_data = response.json()

        if response.ok:
            # Salvar a mensagem enviada no histórico local
            print(f"DEBUG_CALL: Chamando salvar_mensagem para a mensagem enviada de '{numero_cliente_normalizado}'.")
            salvar_mensagem(numero_cliente_normalizado, mensagem_texto, int(time.time()), remetente='sent')

            # Atualiza o status do lead para 'em_conversacao'
            # E REMOVE DOS PENDENTES SE ESTIVER LÁ
            leads = load_leads()
            pending_leads_updated_in_memory = False
            for lead in leads:
                if normalize_phone_number(lead.get('telefone')) == numero_cliente_normalizado: # <<<< NORMALIZAÇÃO AQUI
                    if lead.get('status_contato') not in ["em_conversacao"]:
                        lead['status_contato'] = "em_conversacao"
                        pending_leads_updated_in_memory = True # Indica que o lead foi atualizado em 'leads.json'
                        print(f"DEBUG: Status do lead {numero_cliente_normalizado} atualizado para 'em_conversacao' (via envio de mensagem personalizada).")
                    break
            if pending_leads_updated_in_memory:
                save_leads(leads)

            # <<<< ADICIONADO AQUI: REMOVER LEAD DOS PENDENTES APÓS ENVIO DE MENSAGEM PERSONALIZADA >>>>
            # Isso é crucial para que ele desapareça da lista de pendentes.
            remove_pending_lead_by_phone(numero_cliente_normalizado) # <<<< Usa o número normalizado

            return jsonify(response_data), 200
        else:
            print(f"Erro da API do WhatsApp ao enviar mensagem: {response_data}")
            error_message = response_data.get('error', {}).get('message', 'Erro desconhecido da API do WhatsApp.')
            return jsonify({"erro": error_message}), response.status_code
    except Exception as e:
        print(f"Erro na requisição para a API do WhatsApp (mensagem): {str(e)}")
        return jsonify({"erro": f"Falha interna ao enviar mensagem: {str(e)}"}), 500


# === NOVO ENDPOINT: BOTÕES INTERATIVOS ===
@app.route('/api/enviar-botoes', methods=['POST'])
@login_required
def enviar_botoes_api():
    data = request.json or {}
    numero_original = data.get('numero')
    texto = data.get('texto', 'Deseja falar com um atendente?')
    botoes = data.get('botoes')  # opcional: lista de {"id": "...", "title": "..."}

    numero = normalize_phone_number(numero_original)
    if not numero:
        return jsonify({"erro": "Número inválido após normalização"}), 400

    if not botoes:
        botoes = [
            {"id": "sim_atendente", "title": "Sim"},
            {"id": "nao_atendente", "title": "Não"}
        ]
    if len(botoes) > 3:
        return jsonify({"erro": "Máximo de 3 botões"}), 400

    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "interactive",
        "interactive": {
            "type": "button",
            "body": {"text": texto},
            "action": {
                "buttons": [
                    {"type": "reply", "reply": {"id": b["id"], "title": b["title"]}}
                    for b in botoes
                ]
            }
        }
    }

    try:
        resp = requests.post(url, headers=headers, json=payload)
        data_resp = resp.json()
        if resp.ok:
            salvar_mensagem(numero, f"[BOTÕES] {texto} - Opções: " + ", ".join([b['title'] for b in botoes]), int(time.time()), remetente='sent')
            return jsonify(data_resp), 200
        else:
            msg = data_resp.get('error', {}).get('message', 'Erro desconhecido da API do WhatsApp.')
            return jsonify({"erro": msg}), resp.status_code
    except Exception as e:
        return jsonify({"erro": f"Falha interna ao enviar botões: {str(e)}"}), 500


def carregar_mensagens_lidas():
    mensagens_lidas_file = os.path.join(MESSAGES_DIR, "mensagens_lidas.json")
    if os.path.exists(mensagens_lidas_file):
        try:
            with open(mensagens_lidas_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {}
    return {}


def salvar_mensagens_lidas(dados):
    mensagens_lidas_file = os.path.join(MESSAGES_DIR, "mensagens_lidas.json")
    with open(mensagens_lidas_file, 'w', encoding='utf-8') as f:
        json.dump(dados, f)


@app.route("/api/marcar-como-lido", methods=["POST"])
@login_required
def marcar_como_lido():
    data = request.get_json()
    numero = data.get("numero")
    numero_normalizado = normalize_phone_number(numero)

    if not numero_normalizado:
        return jsonify({"erro": "Número inválido"}), 400

    # Carrega dados existentes
    mensagens_lidas = carregar_mensagens_lidas()

    # Marca como lido com timestamp atual
    mensagens_lidas[numero_normalizado] = int(time.time())

    # Salva no arquivo
    salvar_mensagens_lidas(mensagens_lidas)

    print(f"Marcando mensagens de {numero_normalizado} como lidas")
    return jsonify({"status": "ok"})

# === NOVO ENDPOINT: LISTA INTERATIVA ===
@app.route('/api/enviar-lista', methods=['POST'])
@login_required
def enviar_lista_api():
    data = request.json or {}
    numero_original = data.get('numero')
    numero = normalize_phone_number(numero_original)
    if not numero:
        return jsonify({"erro": "Número inválido após normalização"}), 400

    header_text = data.get('header_text', 'Atendimento')
    body_text = data.get('body_text', 'Escolha uma opção:')
    footer_text = data.get('footer_text', 'Equipe')
    button_text = data.get('button_text', 'Ver opções')
    sections = data.get('sections')

    if not sections or not isinstance(sections, list):
        return jsonify({"erro": "Informe 'sections' como lista"}), 400

    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "interactive",
        "interactive": {
            "type": "list",
            "header": {"type": "text", "text": header_text},
            "body": {"text": body_text},
            "footer": {"text": footer_text},
            "action": {"button": button_text, "sections": sections}
        }
    }

    try:
        resp = requests.post(url, headers=headers, json=payload)
        data_resp = resp.json()
        if resp.ok:
            salvar_mensagem(numero, f"[LISTA] {header_text} — {body_text}", int(time.time()), remetente='sent')
            return jsonify(data_resp), 200
        else:
            msg = data_resp.get('error', {}).get('message', 'Erro desconhecido da API do WhatsApp.')
            return jsonify({"erro": msg}), resp.status_code
    except Exception as e:
        return jsonify({"erro": f"Falha interna ao enviar lista: {str(e)}"}), 500


@app.route('/api/clientes-mensagens', methods=['GET'])
@login_required
def listar_clientes_com_mensagens():
    if not os.path.exists(MESSAGES_DIR):
        return jsonify([])

    numeros_com_mensagens = [
        f.replace('.json', '')
        for f in os.listdir(MESSAGES_DIR)
        if f.endswith('.json')
    ]
    print(f"DEBUG: Clientes com mensagens encontrados: {numeros_com_mensagens}")
    return jsonify(numeros_com_mensagens)


@app.route('/api/numeros', methods=['GET'])
@login_required
def listar_numeros_api():
    try:
        # A sua verificação inicial para a pasta já é boa, mas não trata permissões.
        if not os.path.exists(MESSAGES_DIR):
            return jsonify([])

        numeros = [
            f.replace('.json', '')
            for f in os.listdir(MESSAGES_DIR)
            if f.endswith('.json')
        ]
        print(f"DEBUG: Números para atualização periódica: {numeros}")
        return jsonify(numeros)

    except Exception as e:
        print(f"ERRO CRÍTICO: Falha ao listar números da pasta de mensagens. Erro: {e}")
        # Retorna uma resposta JSON com o código de erro 500 para o frontend.
        return jsonify({"erro": f"Erro interno do servidor: {str(e)}"}), 500


# app.py

@app.route('/api/excluir-contato', methods=['POST'])
@login_required
def excluir_contato_api():
    data = request.json
    numero_original = data.get('numero')

    print(f"DEBUG: Requisição para excluir contato recebida. Número original: {numero_original}")

    numero = normalize_phone_number(numero_original)
    if not numero:
        print(f"DEBUG: Número original '{numero_original}' resultou em normalização inválida.")
        return jsonify({"erro": "Número inválido ou não normalizável para exclusão."}), 400

    print(f"DEBUG: Número normalizado para exclusão: {numero}")

    path = os.path.join(MESSAGES_DIR, f"{numero}.json")
    print(f"DEBUG: Tentando excluir arquivo em: {path}")

    if os.path.exists(path):
        try:
            os.remove(path)
            print(f"DEBUG: Arquivo de chat {path} excluído com sucesso.")

            # ... (o restante do seu código para remover da lista geral de leads) ...

            return jsonify({"mensagem": f"Contato {numero} e histórico de mensagens excluídos."}), 200
        except Exception as e:
            print(f"ERROR: Erro ao excluir arquivo de chat {path}: {e}")
            return jsonify({"erro": f"Falha ao excluir o contato: {str(e)}"}), 500
    else:
        print(f"DEBUG: Arquivo de chat {path} NÃO encontrado. Retornando 404.")
        return jsonify({"mensagem": f"Contato {numero} ou histórico de mensagens não encontrado."}), 404

# Certifique-se de ter essa linha em algum lugar no seu app.py para ver o MESSAGES_DIR na inicialização
print(f"DEBUG: MESSAGES_DIR configurado como: {MESSAGES_DIR}")


### **Novas Rotas de API para Leads Salvos e Pendentes**

@app.route('/api/leads_salvos', methods=['GET'])
@login_required
def get_saved_leads_api():
    """Retorna todos os leads salvos (geral)."""
    leads = load_leads()
    return jsonify(leads)


@app.route('/api/pending_leads', methods=['GET'])
@login_required
def get_pending_leads_api():
    """Retorna todos os leads na fila de espera para contato."""
    pending_leads = load_pending_leads()
    return jsonify(pending_leads)


@app.route('/api/remove_pending_lead', methods=['POST'])
@login_required
def remove_pending_lead_api():
    """Remove um lead específico da fila de espera."""
    data = request.json
    numero_telefone = data.get('telefone')  # Espera o telefone do lead para identificar

    if not numero_telefone:
        return jsonify({"erro": "Telefone do lead não fornecido"}), 400

    if remove_pending_lead_by_phone(numero_telefone):
        return jsonify({"mensagem": "Lead removido da fila de pendentes com sucesso."})
    else:
        return jsonify({"erro": "Lead não encontrado na fila de pendentes ou telefone incorreto."}), 404


try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

@app.route("/api/export", methods=["GET"])
@login_required
def api_export():
    kind   = (request.args.get("kind") or "summary").strip()       # summary|leads|events|report
    fmt    = (request.args.get("format") or "csv").strip().lower() # csv|xlsx
    start  = request.args.get("start")
    end    = request.args.get("end")
    if not start or not end:
        return jsonify({"erro": "Passe start/end no formato YYYY-MM-DD"}), 400

    start_ts = _parse_date_ymd(start)
    end_ts   = _parse_date_ymd(end) + 86399

    rows = []
    filename = f"{kind}_{start}_{end}.{fmt}"

    if kind == "summary":
        data = summarize_metrics(start_ts, end_ts)
        rows = [["métrica", "valor"]] + [[k, str(v)] for k, v in data.items()]

    elif kind == "leads":
        # Cohort: todos os contatos que receberam o 1º envio no período (novos contatos)
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT phone_e164, COALESCE(name,'') AS name,
                   first_contacted_at, initial_replied, outcome, outcome_set_at
            FROM leads
            WHERE first_contacted_at IS NOT NULL
              AND first_contacted_at BETWEEN ? AND ?
            ORDER BY first_contacted_at DESC
        """, (start_ts, end_ts))
        leads_rows = cur.fetchall()
        conn.close()

        # Resumo desta coorte
        total       = len(leads_rows)  # "Contatos novos no mês"
        responded   = sum(1 for r in leads_rows if r["initial_replied"] == 1)
        no_response = total - responded
        won         = sum(1 for r in leads_rows if (r["outcome"] or "") == "won")
        lost        = sum(1 for r in leads_rows if (r["outcome"] or "") == "lost")

        # >>> Listagem sem colunas de outcome <<<
        headers = ["telefone","nome","primeiro_contato_ts","primeiro_contato_data","respondeu_inicial"]

        def _row_to_list_simple(r):
            ts1 = r["first_contacted_at"]
            return [
                r["phone_e164"],
                r["name"],
                ts1,
                (datetime.fromtimestamp(ts1).strftime("%Y-%m-%d %H:%M:%S") if ts1 else ""),
                r["initial_replied"],
            ]

        if fmt == "xlsx":
            if not HAS_OPENPYXL:
                return jsonify({"erro": "openpyxl não instalado. Use format=csv ou instale openpyxl."}), 400
            wb = Workbook()

            # Aba Resumo
            ws1 = wb.active; ws1.title = "Resumo"
            ws1.append(["Período", f"{start} → {end}"])
            ws1.append(["Contatos novos no mês", total])
            ws1.append(["Responderam",            responded])
            ws1.append(["Não responderam",        no_response])
            ws1.append(["Fecharam (won)",         won])
            ws1.append(["Não fecharam (lost)",    lost])

            # Aba com "Novos contatos" (lista)
            ws2 = wb.create_sheet(title="Novos_contatos")
            ws2.append(headers)
            for r in leads_rows:
                ws2.append(_row_to_list_simple(r))

            out = io.BytesIO()
            wb.save(out); out.seek(0)
            resp = Response(out.getvalue(),
                            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp.headers["Content-Disposition"] = f'attachment; filename="leads_{start}_{end}.xlsx"'
            return resp

        else:
            # CSV: Resumo no topo + linha em branco + Listagem simples (sem outcome)
            buf = io.StringIO(); w = csv.writer(buf)
            w.writerow(["Período", f"{start} → {end}"])
            w.writerow(["Contatos novos no mês", total])
            w.writerow(["Responderam",            responded])
            w.writerow(["Não responderam",        no_response])
            w.writerow(["Fecharam (won)",         won])
            w.writerow(["Não fecharam (lost)",    lost])
            w.writerow([])

            w.writerow(headers)
            for r in leads_rows:
                w.writerow(_row_to_list_simple(r))

            resp = Response(buf.getvalue(), mimetype="text/csv; charset=utf-8")
            resp.headers["Content-Disposition"] = f'attachment; filename="leads_{start}_{end}.csv"'
            return resp

    elif kind == "events":
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT phone_e164, type, ts, meta
            FROM events
            WHERE ts BETWEEN ? AND ?
            ORDER BY ts DESC
        """, (start_ts, end_ts))
        rows = [["telefone","tipo","timestamp","meta_json"]]
        for r in cur.fetchall():
            rows.append([r["phone_e164"], r["type"], r["ts"], r["meta"] or ""])
        conn.close()



    elif kind == "full":

        # Cohort: todos que receberam o 1º envio no período (novos contatos)

        conn = get_db();
        cur = conn.cursor()

        cur.execute("""

            SELECT phone_e164, COALESCE(name,'') AS name,

                   first_contacted_at, initial_replied, outcome, outcome_set_at

            FROM leads

            WHERE first_contacted_at IS NOT NULL

              AND first_contacted_at BETWEEN ? AND ?

            ORDER BY first_contacted_at DESC

        """, (start_ts, end_ts))

        leads_rows = cur.fetchall()

        conn.close()

        # Resumo desta coorte (para cards)

        total = len(leads_rows)

        responded = sum(1 for r in leads_rows if r["initial_replied"] == 1)

        no_response = total - responded

        won = sum(1 for r in leads_rows if (r["outcome"] or "") == "won")

        lost = sum(1 for r in leads_rows if (r["outcome"] or "") == "lost")

        # >>> LISTAGEM COMPLETA SEM COLUNAS DE OUTCOME <<<

        headers_full = [

            "telefone", "nome",

            "primeiro_contato_ts", "primeiro_contato_data",

            "respondeu_inicial"

        ]

        def _row_full(r):

            ts1 = r["first_contacted_at"]

            return [

                r["phone_e164"],

                r["name"],

                ts1,

                (datetime.fromtimestamp(ts1).strftime("%Y-%m-%d %H:%M:%S") if ts1 else ""),

                r["initial_replied"],

            ]

        if fmt == "xlsx":

            if not HAS_OPENPYXL:
                return jsonify({"erro": "openpyxl não instalado. Use format=csv ou instale openpyxl."}), 400

            wb = Workbook()

            # Aba Resumo (só números agregados)

            ws1 = wb.active;
            ws1.title = "Resumo"

            ws1.append(["Período", f"{start} → {end}"])

            ws1.append(["Contatos novos no mês", total])

            ws1.append(["Responderam", responded])

            ws1.append(["Não responderam", no_response])

            ws1.append(["Fecharam (won)", won])

            ws1.append(["Não fecharam (lost)", lost])

            # Aba com a listagem SEM outcome

            ws2 = wb.create_sheet(title="Novos_contatos")

            ws2.append(headers_full)

            for r in leads_rows:
                ws2.append(_row_full(r))

            out = io.BytesIO()

            wb.save(out);
            out.seek(0)

            resp = Response(out.getvalue(),

                            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            resp.headers["Content-Disposition"] = f'attachment; filename="full_{start}_{end}.xlsx"'

            return resp


        else:

            # CSV: resumo no topo + linha em branco + listagem SEM outcome

            buf = io.StringIO();
            w = csv.writer(buf)

            w.writerow(["Período", f"{start} → {end}"])

            w.writerow(["Contatos novos no mês", total])

            w.writerow(["Responderam", responded])

            w.writerow(["Não responderam", no_response])

            w.writerow(["Fecharam (won)", won])

            w.writerow(["Não fecharam (lost)", lost])

            w.writerow([])

            w.writerow(headers_full)

            for r in leads_rows:
                w.writerow(_row_full(r))

            resp = Response(buf.getvalue(), mimetype="text/csv; charset=utf-8")

            resp.headers["Content-Disposition"] = f'attachment; filename="full_{start}_{end}.csv"'

            return resp



    else:
        return jsonify({"erro": "kind inválido (use summary|leads|events|report)"}), 400

    # Saída para summary/events (quando não retornamos acima)
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        for row in rows:
            w.writerow(row)
        resp = Response(buf.getvalue(), mimetype="text/csv; charset=utf-8")
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    elif fmt == "xlsx":
        if not HAS_OPENPYXL:
            return jsonify({"erro": "openpyxl não instalado. Use format=csv ou instale openpyxl."}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = kind
        for row in rows:
            ws.append(row)
        out = io.BytesIO()
        wb.save(out); out.seek(0)
        resp = Response(out.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    else:
        return jsonify({"erro": "format inválido (use csv|xlsx)"}), 400

def _month_range_previous():
    """Retorna (start_date, end_date) do mês anterior em YYYY-MM-DD."""
    today = date.today()
    first_this = date(today.year, today.month, 1)
    last_prev  = first_this - timedelta(days=1)
    first_prev = date(last_prev.year, last_prev.month, 1)
    start = first_prev.strftime("%Y-%m-%d")
    end   = last_prev.strftime("%Y-%m-%d")
    return start, end

def _parse_date_ymd(s: str) -> int:
    # (se já existir no seu projeto, reutilize a existente)
    return int(datetime.strptime(s, "%Y-%m-%d").timestamp())

def _build_leads_export_file(start:str, end:str, fmt:str="xlsx"):
    """
    Gera o arquivo da opção 'leads do período' (Resumo + Lista sem outcome),
    salva em /data/exports e retorna (filepath, filename, mime).
    """
    start_ts = _parse_date_ymd(start)
    end_ts   = _parse_date_ymd(end) + 86399

    # Query dos 'novos contatos no período' (first_contacted_at dentro do range)
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT phone_e164, COALESCE(name,'') AS name,
               first_contacted_at, initial_replied, outcome, outcome_set_at
        FROM leads
        WHERE first_contacted_at IS NOT NULL
          AND first_contacted_at BETWEEN ? AND ?
        ORDER BY first_contacted_at DESC
    """, (start_ts, end_ts))
    leads_rows = cur.fetchall()
    conn.close()

    # Resumo da coorte
    total       = len(leads_rows)
    responded   = sum(1 for r in leads_rows if r["initial_replied"] == 1)
    no_response = total - responded
    won         = sum(1 for r in leads_rows if (r["outcome"] or "") == "won")
    lost        = sum(1 for r in leads_rows if (r["outcome"] or "") == "lost")

    headers = ["telefone","nome","primeiro_contato_ts","primeiro_contato_data","respondeu_inicial"]
    def _row_simple(r):
        ts1 = r["first_contacted_at"]
        return [
            r["phone_e164"],
            r["name"],
            ts1,
            (datetime.fromtimestamp(ts1).strftime("%Y-%m-%d %H:%M:%S") if ts1 else ""),
            r["initial_replied"],
        ]

    # Monta nome do arquivo
    base = f"leads_{start}_{end}"
    if fmt == "xlsx":
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl não instalado (use CSV ou instale openpyxl).")
        filename = base + ".xlsx"
        filepath = os.path.join(EXPORTS_DIR, filename)

        wb = Workbook()
        ws1 = wb.active; ws1.title = "Resumo"
        ws1.append(["Período", f"{start} → {end}"])
        ws1.append(["Contatos novos no mês", total])
        ws1.append(["Responderam",            responded])
        ws1.append(["Não responderam",        no_response])
        ws1.append(["Fecharam (won)",         won])
        ws1.append(["Não fecharam (lost)",    lost])

        ws2 = wb.create_sheet(title="Novos_contatos")
        ws2.append(headers)
        for r in leads_rows:
            ws2.append(_row_simple(r))

        wb.save(filepath)
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return filepath, filename, mime

    # CSV
    filename = base + ".csv"
    filepath = os.path.join(EXPORTS_DIR, filename)
    with open(filepath, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Período", f"{start} → {end}"])
        w.writerow(["Contatos novos no mês", total])
        w.writerow(["Responderam",            responded])
        w.writerow(["Não responderam",        no_response])
        w.writerow(["Fecharam (won)",         won])
        w.writerow(["Não fecharam (lost)",    lost])
        w.writerow([])
        w.writerow(headers)
        for r in leads_rows:
            w.writerow(_row_simple(r))
    mime = "text/csv"
    return filepath, filename, mime


def whatsapp_upload_document(filepath: str, mime: str) -> str | None:
    """Sobe o arquivo para a mídia do WhatsApp e retorna media_id."""
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/media"
    headers = { "Authorization": f"Bearer {ACCESS_TOKEN}" }
    with open(filepath, "rb") as f:
        files = { "file": (os.path.basename(filepath), f, mime) }
        data  = { "messaging_product": "whatsapp" }
        r = requests.post(url, headers=headers, files=files, data=data, timeout=120)
    try:
        j = r.json()
    except Exception:
        j = {}
    if r.ok and j.get("id"):
        return j["id"]
    print("ERROR upload media:", r.status_code, j)
    return None

def whatsapp_send_document(to: str, media_id: str, filename: str):
    """Envia a mensagem de documento usando o media_id já subido."""
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": normalize_phone_number(to),
        "type": "document",
        "document": {
            "id": media_id,
            "filename": filename
        }
    }
    r = requests.post(url, headers=headers, json=payload, timeout=60)
    try:
        j = r.json()
    except Exception:
        j = {}
    if not r.ok:
        print("ERROR send document:", r.status_code, j)
    else:
        print("OK document sent:", j)
def job_send_prev_month_leads(fmt: str = "xlsx"):
    # CSV não é aceito pelo WhatsApp -> força XLSX
    if fmt.lower() == "csv":
        fmt = "xlsx"

    if not ATTENDANT_NUMBER:
        print("ATTENTION: defina ATTENDANT_NUMBER no .env")
        return

    start, end = _month_range_previous()
    try:
        filepath, filename, mime = _build_leads_export_file(start, end, fmt=fmt)
        media_id = whatsapp_upload_document(filepath, mime)
        if not media_id:
            return
        whatsapp_send_document(ATTENDANT_NUMBER, media_id, filename)
        print(f"Monthly export enviado ao atendente: {filename}")
    except Exception as e:
        print("ERROR monthly export:", e)


# Agenda: todo dia 1 às 08:00 (timezone do servidor)
scheduler = BackgroundScheduler()
scheduler.add_job(
    func=job_send_prev_month_leads,
    trigger="cron",
    day=1, hour=8, minute=0,
    id="monthly_leads_export",
    replace_existing=True
)
# Evita duplicar em modo debug com reloader
if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or not app.debug:
    scheduler.start()

# Endpoint opcional para disparar manualmente
@app.route("/api/metrics/send-monthly-export", methods=["POST"])
@login_required
def api_send_monthly_export():
    body = request.get_json(silent=True) or {}
    fmt  = (body.get("format") or "xlsx").lower()

    # CSV não é aceito pelo WhatsApp -> força XLSX
    if fmt == "csv":
        fmt = "xlsx"

    month = (body.get("month") or "").strip()
    if month:
        y, m = map(int, month.split("-"))
        first = date(y, m, 1)
        last  = date(y, m+1, 1) - timedelta(days=1) if m < 12 else date(y, 12, 31)
        start = first.strftime("%Y-%m-%d"); end = last.strftime("%Y-%m-%d")
    else:
        start, end = _month_range_previous()

    try:
        filepath, filename, mime = _build_leads_export_file(start, end, fmt=fmt)
        media_id = whatsapp_upload_document(filepath, mime)
        if not media_id:
            return jsonify({"ok": False, "erro": "falha no upload"}), 500
        whatsapp_send_document(ATTENDANT_NUMBER, media_id, filename)
        return jsonify({"ok": True, "filename": filename, "range": {"start": start, "end": end}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500



if __name__ == "__main__":
    # Garante que o arquivo de usuários existe e tem um usuário padrão
    load_users()
    # Chama a função para garantir que todos os leads existentes tenham um status
    migrate_leads_status()
    app.run(host="0.0.0.0", port=5000, debug=True)